import urllib.request
import json
import uuid
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def send_telegram_message(bot_token: str, chat_id: str, text: str) -> bool:
    if not bot_token or not chat_id:
        return False
        
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {
        'chat_id': chat_id,
        'text': text,
        'parse_mode': 'HTML',
        'disable_web_page_preview': True
    }
    
    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers={'Content-Type': 'application/json'})
    
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            res = json.loads(resp.read().decode('utf-8'))
            return res.get('ok', False)
    except Exception as e:
        print(f"Telegram Mesaj Hatası: {e}")
        return False

def send_telegram_document(bot_token: str, chat_id: str, file_bytes: bytes, filename: str, caption: str = "") -> bool:
    if not bot_token or not chat_id:
        return False
        
    boundary = uuid.uuid4().hex
    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    
    body = []
    
    # chat_id
    body.append(f'--{boundary}'.encode('utf-8'))
    body.append(f'Content-Disposition: form-data; name="chat_id"'.encode('utf-8'))
    body.append(b'')
    body.append(str(chat_id).encode('utf-8'))
    
    # caption
    if caption:
        body.append(f'--{boundary}'.encode('utf-8'))
        body.append(f'Content-Disposition: form-data; name="caption"'.encode('utf-8'))
        body.append(b'')
        body.append(caption.encode('utf-8'))
        
    # document
    body.append(f'--{boundary}'.encode('utf-8'))
    body.append(f'Content-Disposition: form-data; name="document"; filename="{filename}"'.encode('utf-8'))
    body.append(b'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    body.append(b'')
    body.append(file_bytes)
    
    body.append(f'--{boundary}--'.encode('utf-8'))
    body.append(b'')
    
    payload_data = b'\r\n'.join(body)
    
    headers = {
        'Content-Type': f'multipart/form-data; boundary={boundary}',
        'Content-Length': str(len(payload_data))
    }
    
    req = urllib.request.Request(url, data=payload_data, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            res = json.loads(resp.read().decode('utf-8'))
            return res.get('ok', False)
    except Exception as e:
        print(f"Telegram Doküman Hatası: {e}")
        return False

def generate_excel_bulletin(matches: list, title: str = "SULEYMANDO BÜLTENİ") -> bytes:
    """
    Formüle uyan maçları renkli ve biçimlendirilmiş Excel (.xlsx) dosyasına dönüştürür.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formüle Uyan Maçlar"
    
    # Header Styling
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "Tarih", "Saat", "Kod", "MBS", "Ev Sahibi", "Deplasman", 
        "MS 1", "MS X", "MS 2", "Favori Taraf", "Favori Oran", 
        "İY 1.5 Üst Oran", "Ana Tahmin", "Ekstra Tahmin", "İY Skoru", "MS Skoru", "İY 1.5 Üst Durumu"
    ]
    
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # Row Data
    for m in matches:
        is_home_fav = (m['fav_side'] == 'EV SAHİBİ')
        iy_ust_odd = m.get('iy_1_5_ust')
        iy_ust_str = f"{iy_ust_odd:.2f}" if iy_ust_odd else "N/A"
        
        status_str = m.get('iy_1_5_status', 'OYNANMADI')
        if status_str == 'TUTTU':
            status_display = "✅ TUTTU"
        elif status_str == 'YATTI':
            status_display = "❌ YATTI"
        else:
            status_display = "⏳ OYNANMADI"
            
        row_values = [
            m.get('date', ''),
            m.get('time', ''),
            m.get('code', ''),
            m.get('mbs', 1),
            m.get('home', ''),
            m.get('away', ''),
            m.get('ms1', 0.0),
            m.get('msx', 0.0),
            m.get('ms2', 0.0),
            m.get('fav_side', ''),
            m.get('fav_odds', 0.0),
            iy_ust_str,
            "İY 1.5 ÜST",
            m.get('extra_prediction', ''),
            m.get('iy_score', '-'),
            m.get('ms_score', '-'),
            status_display
        ]
        ws.append(row_values)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def format_telegram_2h_bulletin(matches: list) -> str:
    """
    Önümüzdeki 2 saat içinde başlayacak maçların kısa ve öz Telegram mesaj şablonu (Spam önleyici).
    """
    if not matches:
        return "⏳ <b>SULEYMANDO BİLDİRİMİ</b>\n\nÖnümüzdeki 2 saat içinde başlayacak formüle uyan maç bulunmamaktadır."
        
    msg = f"⏳ <b>SULEYMANDO YAKLAŞAN 2 SAATLİK MAÇLAR</b>\n"
    msg += f"📊 <b>Başlayacak Maç Sayısı: {len(matches)}</b>\n"
    msg += "━━━━━━━━━━━━━━━━━━━\n\n"
    
    for idx, m in enumerate(matches, 1):
        iy_ust_odd = m.get('iy_1_5_ust')
        iy_ust_str = f"{iy_ust_odd:.2f}" if iy_ust_odd else "N/A"
        fav_icon = "👑 EV" if m['fav_side'] == 'EV SAHİBİ' else "✈️ DEP"
        
        msg += (
            f"<b>{idx}. ⏰ {m['time']}</b> | <code>MBS:{m.get('mbs',1)}</code> | 📌 <code>{m['code']}</code>\n"
            f"⚽ <b>{m['home']}</b> vs <b>{m['away']}</b> (Fav: {fav_icon})\n"
            f"🔥 <b>İY 1.5 ÜST:</b> <code>{iy_ust_str}</code> | 🎯 <b>Ekstra:</b> {m['extra_prediction']}\n"
            f"───────────────────\n"
        )
        
    return msg

def format_telegram_winrate_report(matches: list, date_str: str) -> str:
    """
    Gece 23:45 Gün Sonu Başarı ve İY Skor Doğrulama Raporu.
    """
    played_matches = [m for m in matches if m.get('iy_1_5_status') in ['TUTTU', 'YATTI']]
    won_matches = [m for m in played_matches if m.get('iy_1_5_status') == 'TUTTU']
    lost_matches = [m for m in played_matches if m.get('iy_1_5_status') == 'YATTI']
    
    total_count = len(played_matches)
    won_count = len(won_matches)
    lost_count = len(lost_matches)
    
    win_rate = (won_count / total_count * 100) if total_count > 0 else 0.0
    
    report = (
        f"🏆 <b>SULEYMANDO GÜN SONU BAŞARI RAPORU (23:45)</b>\n"
        f"📅 <b>Tarih: {date_str}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━\n\n"
        f"✅ <b>Tutan İY 1.5 ÜST Maç Sayısı:</b> {won_count}\n"
        f"❌ <b>Yatan Maç Sayısı:</b> {lost_count}\n"
        f"📊 <b>Oynanan Toplam Maç:</b> {total_count}\n"
        f"🔥 <b>GÜNÜN BAŞARI ORANI:</b> <b>%{win_rate:.1f}</b>\n\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"📎 <i>Detaylı maç listesi ve İY skor doğrulama raporu ekteki Excel dosyasındadır.</i>"
    )
    return report
